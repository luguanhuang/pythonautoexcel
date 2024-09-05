
from PyQt5.QtWidgets import *
from PyQt5.QtMultimedia import *
from PyQt5.QtWidgets import  QApplication,QFileDialog
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from Ui_erp import Ui_MainWindow
import sys
# from services.sport.sportquery
from dao import erpinfo
from PyQt5.QtWidgets import QApplication, QMainWindow
from openpyxl.styles import Font, colors
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import colors
import openpyxl
from openpyxl.styles import Border, Side
import pymssql
from pymssql import _mssql
from pymssql import _pymssql
import uuid
import decimal
from config import setting 
from utils.log import erplogger
import threading

# class onerowdata:
#     def __init__(self, row):
#         self.mainnum = row[0]
#         self.mainname = row[1]
#         self.mainspec = row[2]
#     mainnum = ""
#     mainname = ""
#     mainspec = ""
# class mainpartinfo:
#     def __init__(self, subpartnum, subpartname, row):
#         self.subpartnum = subpartnum
#         self.subpartname = subpartname
#         self.arrmainpartinfo.clear()
#         datainfo = onerowdata(row)

#         erplogger.info(f"mainnum1111111112222222222={row[0]}")
#         # self.arrmainpartinfo.append(maininfo)
#         self.arrmainpartinfo.append(datainfo)
#     def display(self):
#         for datainfo in self.arrmainpartinfo:
#                 # erplogger.info(f"mainnum={datainfo['mainnum']}, mainname={datainfo['mainname']}, mainspec={datainfo['mainspec']}, mainsize={datainfo['mainsize']}, mainnotes={datainfo['mainnotes']}, dienum={datainfo['dienum']}, prodnum={datainfo['prodnum']}, prodcnt={datainfo['prodcnt']}, subpartnum={datainfo['subpartnum']}, subpartname={datainfo['subpartname']}")
#             # erplogger.info(f"mainnum11={datainfo.mainnum}, mainname22={datainfo.mainname}, mainspec33={datainfo.mainspec}")
#             print(f"mainnum11={datainfo.mainnum}, mainname22={datainfo.mainname}, mainspec33={datainfo.mainspec}")
#                 # erplogger.info(f"subpartnum2222={data['subpartnum']}, subpartname={data['subpartname']}")
            
#     arrmainpartinfo = []
#     subpartnum = ""
#     subpartname = ""

class sportMainWindow(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super(Ui_MainWindow, self).__init__()
        self.setupUi(self)
        self.destpath = "D:/sample.xlsx"
        self.editFilePath.setText(self.destpath)
        self.buttonnexportfile.clicked.connect(self.exportfile)   # 打开视频文件按钮

        self.selectfilepath.clicked.connect(self.filepathselect)   # 打开视频文件按钮
        
        self.dateTimeEdit.setCalendarPopup(True)
        
        self.dateTimeEdit.setDisplayFormat("yyyy-MM-dd")
        self.dateTimeEdit.setDateTime(QDateTime.currentDateTime())

    def filepathselect(self):
        print("filepathselect: func begin")
        
        self.destpath, filetype = QFileDialog.getSaveFileName(self, "文件保存", "myfile.xlsx", "Excel工作簿 (*.xlsx)")
        print("destpath=", self.destpath)
        

        if self.destpath:  # 如果获取的路径非空
            self.editFilePath.setText(self.destpath)

    def setsheetwidth(self, ws):
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 7
        ws.column_dimensions['I'].width = 7
        ws.column_dimensions['J'].width = 4
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12

    def DisTaskOrderInfo(self, ws, onepagebegin, pagenum):
        ws.merge_cells(start_row=onepagebegin, end_row=onepagebegin+1, start_column=3, end_column=9)#标底合并
        ws['C'+str(onepagebegin)] = '生产任务单'
        align = Alignment(horizontal='center', vertical='center')
        ws['C'+str(onepagebegin)].alignment = align

        cell = ws["C"+str(onepagebegin)]
        font = Font(color=colors.BLACK, bold=True, size=20)
        cell.font = font

        # ws.merge_cells(start_row=onepagebegin, end_row=onepagebegin, start_column=onepagebegin, end_column=2)#标底合并
        ws.merge_cells(start_row=onepagebegin, end_row=onepagebegin, start_column=1, end_column=2)#标底合并
        
        ws["A"+str(onepagebegin)] = "PEPSEN"
        cell = ws["A"+str(onepagebegin)]
        font = Font(color=colors.BLUE, size = 10, bold=True)
        cell.font = font

        font = Font(color=colors.BLACK, bold=True, size=10)
        # ws.merge_cells(start_row=onepagebegin+1, end_row=onepagebegin+1, start_column=1, end_column=onepagebegin+1)#标底合并
        ws.merge_cells(start_row=onepagebegin+1, end_row=onepagebegin+1, start_column=1, end_column=2)#标底合并
        cell = ws["A"+str(onepagebegin+1)]
        cell.font = font
        ws["A"+str(onepagebegin+1)] = "班组:绕注乙班"
        ws['M'+str(onepagebegin)] = str(pagenum)
        return onepagebegin+3

    def DisMainPartHeader(self, ws, onepagebegin):
        align = Alignment(horizontal='center', vertical='center')
        font = Font(color=colors.BLACK, bold=True, size=10)
        fieldcolor = 'D200D2'
        nextbegin = str(onepagebegin)
        ws["A"+(nextbegin)] = "主件品号"
        font = Font(color=fieldcolor, bold=True, size = 10)
        # ws["A4"].font = font
        # ws['A4'].alignment = align
        ws["A"+(nextbegin)].font = font
        ws['A'+(nextbegin)].alignment = align
        # tmpcell = 'B4:C4'
        tmpcell = 'B'+(nextbegin)+":C"+(nextbegin)
        ws.merge_cells(tmpcell)
        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['B'+(nextbegin)].fill = fillinfo
        ws["B"+(nextbegin)] = "主件品名"
        ws["B"+(nextbegin)].font = font
        ws['B'+(nextbegin)].alignment = align
        # ws["B4"].border = border
        # tmpcell = 'D4:E4'
        tmpcell = 'D'+(nextbegin)+":E"+(nextbegin)
        ws.merge_cells(tmpcell)
        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['D'+(nextbegin)].fill = fillinfo
        ws["D"+(nextbegin)] = "主件规格"
        ws["D"+(nextbegin)].font = font
        ws['D'+(nextbegin)].alignment = align
        # ws["D4"].border = border

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['F'+(nextbegin)].fill = fillinfo
        ws["F"+(nextbegin)] = "模具尺寸"
        ws["F"+(nextbegin)].font = font
        ws['F'+(nextbegin)].alignment = align
        # ws["F4"].border = border

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['G'+(nextbegin)].fill = fillinfo
        ws["G"+(nextbegin)] = "单头备注"
        ws["G"+(nextbegin)].font = font
        ws['G'+(nextbegin)].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['H'+(nextbegin)].fill = fillinfo
        ws["H"+(nextbegin)] = "模具数量"
        ws["H"+(nextbegin)].font = font
        ws['H'+(nextbegin)].alignment = align
        

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['I'+(nextbegin)].fill = fillinfo
        ws["I"+(nextbegin)] = "生产次数"
        ws["I"+(nextbegin)].font = font
        ws['I'+(nextbegin)].alignment = align
        

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['J'+(nextbegin)].fill = fillinfo
        # tmpcell = 'J4:K4'
        tmpcell = 'J'+(nextbegin)+":K"+(nextbegin)
        ws.merge_cells(tmpcell)
        ws["J"+(nextbegin)] = "生产数量"
        ws["J"+(nextbegin)].font = font
        ws['J'+(nextbegin)].alignment = align
        

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['L'+(nextbegin)].fill = fillinfo
        ws["L"+(nextbegin)] = "单次用量(kg)"
        ws["L"+(nextbegin)].font = font
        ws['L'+(nextbegin)].alignment = align
        

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['M'+(nextbegin)].fill = fillinfo
        ws["M"+(nextbegin)] = "合计用量(kg)"
        ws["M"+(nextbegin)].font = font
        ws['M'+(nextbegin)].alignment = align

    def DisMainPartBody(self, ws, onepagebegin, valueinfo):
        align = Alignment(horizontal='center', vertical='center')
        font1 = Font(size = 9)
        nextbegin = str(onepagebegin)
        # ws["A"+nextbegin] = "B090800001"
        ws["A"+nextbegin] = valueinfo["mainnum"]
        ws["A"+nextbegin].font = font1
        fillinfo = openpyxl.styles.PatternFill(start_color='F9F900', end_color='F9F900', fill_type='solid')
        ws['A'+nextbegin].fill = fillinfo

        # ws["B"+nextbegin] = "交通桩/PU1"
        ws["B"+nextbegin] = valueinfo["mainname"]
        tmpcell = 'B'+nextbegin+":C"+nextbegin
        # ws.merge_cells('B5:C5')
        ws.merge_cells(tmpcell)
        ws["B"+nextbegin].font = font1
        
        # ws["D"+nextbegin] = "270*172/深灰色/95-98A"
        ws["D"+nextbegin] = valueinfo['mainspec']
        tmpcell = 'D'+nextbegin+":E"+nextbegin
        ws.merge_cells(tmpcell)
        ws["D"+nextbegin].font = font1
        # ws["F"+nextbegin] = "270*172专-2"
        ws["F"+nextbegin] = valueinfo['mainsize']
        ws["F"+nextbegin].font = font1
       
        font2 = Font(bold=True, size = 9)
        ws["G"+nextbegin] = valueinfo['mainnotes']
        # ws["G"+nextbegin] = "下单1000件"
        ws["G"+nextbegin].font = font2

        # ws["H"+nextbegin] = "2"
        ws["H"+nextbegin] = valueinfo['dienum']
        ws["H"+nextbegin].font = font1
        ws['H'+nextbegin].alignment = align
        # ws["I"+nextbegin] = "16"
        ws["I"+nextbegin] = valueinfo['prodcnt']
        ws["I"+nextbegin].font = font2
        ws['I'+nextbegin].alignment = align

        tmpcell = 'J'+nextbegin+":K"+nextbegin
        ws.merge_cells(tmpcell)
        ws["J"+nextbegin] = valueinfo['prodnum']
        # ws["J"+nextbegin] = "32"
        ws["J"+nextbegin].font = font1
        ws['J'+nextbegin].alignment = align
        ws["L"+nextbegin] = valueinfo["onecnt"]
        ws["L"+nextbegin].font = font1
        ws['L'+nextbegin].alignment = align

        ws["M"+nextbegin] = valueinfo["totalcnt"]
        ws["M"+nextbegin].font = font1
        ws['M'+nextbegin].alignment = align

    def DisSubPar(self, ws, onepagebegin, arrsubpart):
         
        font1 = Font(size = 9)
        fieldcolor = 'D200D2'
        font = Font(color=fieldcolor, bold=True, size = 10)
        
        nextbegin = str(onepagebegin)
        align = Alignment(horizontal='center', vertical='center')
        ws["A"+nextbegin] = "子件品号"
        ws["A"+nextbegin].font = font
        ws['A'+nextbegin].alignment = align
        tmpcell = 'H'+nextbegin+":I"+nextbegin
        ws.merge_cells(tmpcell)
        
        beginc = "B"
        for data in arrsubpart:
            # print("subpartdata=", data['subpartnum'])   
           
            ws[beginc+nextbegin] = data['subpartnum']
            ws[beginc+nextbegin].font = font1
            ws[beginc+nextbegin].alignment = align
            beginc = chr(ord(beginc)+1)
        # ws["B"+nextbegin] = "D010100009"
        # ws["B"+nextbegin].font = font1
        # ws['B'+nextbegin].alignment = align
        # ws["C"+nextbegin] = "C020600006"
        # ws["C"+nextbegin].font = font1
        # ws['C'+nextbegin].alignment = align
        # ws["D"+nextbegin] = "C020800010"
        # ws["D"+nextbegin].font = font1
        # ws['D'+nextbegin].alignment = align
        # ws["E"+nextbegin] = "0"
        # ws["E"+nextbegin].font = font1
        # ws['E'+nextbegin].alignment = align
        # ws["F"+nextbegin] = "0"
        # ws["F"+nextbegin].font = font1
        # ws['F'+nextbegin].alignment = align
        # ws["G"+nextbegin] = "0"
        # ws["G"+nextbegin].font = font1
        # ws['G'+nextbegin].alignment = align

       
        # ws["H"+nextbegin] = "0"
        # ws["H"+nextbegin].font = font1
        # ws['H'+nextbegin].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['J'+nextbegin].fill = fillinfo
        ws["J"+nextbegin] = "班次"
        ws["J"+nextbegin].font = font
        ws['J'+nextbegin].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['K'+nextbegin].fill = fillinfo
        ws["K"+nextbegin] = "设备号"
        ws["K"+nextbegin].font = font
        ws['K'+nextbegin].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['L'+nextbegin].fill = fillinfo
        ws["L"+nextbegin] = "操作人"
        ws["L"+nextbegin].font = font
        ws['L'+nextbegin].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['M'+nextbegin].fill = fillinfo
        ws["M"+nextbegin] = "设备温度"
        ws["M"+nextbegin].font = font
        ws['M'+nextbegin].alignment = align

        nextbegin = str(onepagebegin+1)
        ws["A"+nextbegin] = "子件品名"
        ws["A"+nextbegin].font = font
        ws['A'+nextbegin].alignment = align

        tmpcell = 'H'+nextbegin+":I"+nextbegin
        ws.merge_cells(tmpcell)

        beginc = "B"
        for data in arrsubpart:
            # print("subpartdata=", data['subpartname'])   
           
            ws[beginc+nextbegin] = data['subpartname']
            ws[beginc+nextbegin].font = font1
            ws[beginc+nextbegin].alignment = align
            beginc = chr(ord(beginc)+1)

        # ws["B"+nextbegin] = "PU1-55D-6.8%"
        # ws["B"+nextbegin].font = font1
        # ws['B'+nextbegin].alignment = align
        # ws["C"+nextbegin] = "MOCA（崇舜）"
        # ws["C"+nextbegin].font = font1
        # ws['C'+nextbegin].alignment = align
        # ws["D"+nextbegin] = "深灰色浆"
        # ws["D"+nextbegin].font = font1
        # ws['D'+nextbegin].alignment = align
        # ws["E"+nextbegin] = "0"
        # ws["E"+nextbegin].font = font1
        # ws['E'+nextbegin].alignment = align
        # ws["F"+nextbegin] = "0"
        # ws["F"+nextbegin].font = font1
        # ws['F'+nextbegin].alignment = align
        # ws["G"+nextbegin] = "0"
        # ws["G"+nextbegin].font = font1
        # ws['G'+nextbegin].alignment = align

        
        # ws["H"+nextbegin] = "0"
        # ws["H"+nextbegin].font = font1
        # ws['H'+nextbegin].alignment = align

        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['J'+nextbegin].fill = fillinfo
        ws["J"+nextbegin] = "早"
        ws["J"+nextbegin].font = font
        ws['J'+nextbegin].alignment = align

        # ws["K"+nextbegin] = "P-045"
        # ws["K"+nextbegin].font = font1
        # ws['K'+nextbegin].alignment = align


        # ws["M"+nextbegin] = "120℃"
        # ws["M"+nextbegin].font = font1
        # ws['M'+nextbegin].alignment = align

        nextbegin = str(onepagebegin+2)
        
        ws["A"+nextbegin] = "配料区（g）"
        ws["A"+nextbegin].font = font
        ws['A'+nextbegin].alignment = align

        beginc = "B"
        for data in arrsubpart:
            ws[beginc+nextbegin] = data['IngredientsArea'] *1000
            ws[beginc+nextbegin].font = font1
            ws[beginc+nextbegin].alignment = align
            beginc = chr(ord(beginc)+1)

        # ws["B"+nextbegin] = "2250"
        # ws["B"+nextbegin].font = font1
        # ws['B'+nextbegin].alignment = align
        # ws["C"+nextbegin] = "448"
        # ws["C"+nextbegin].font = font1
        # ws['C'+nextbegin].alignment = align
        # ws["D"+nextbegin] = "22.6"
        # ws["D"+nextbegin].font = font1
        # ws['D'+nextbegin].alignment = align
        # ws["E"+nextbegin] = "0"
        # ws["E"+nextbegin].font = font1
        # ws['E'+nextbegin].alignment = align
        # ws["F"+nextbegin] = "0"
        # ws["F"+nextbegin].font = font1
        # ws['F'+nextbegin].alignment = align
        # ws["G"+nextbegin] = "0"
        # ws["G"+nextbegin].font = font1
        # ws['G'+nextbegin].alignment = align

        tmpcell = 'H'+nextbegin+":I"+nextbegin
        ws.merge_cells(tmpcell)
        
        fillinfo = openpyxl.styles.PatternFill(start_color='D1EEEE', end_color='D1EEEE', fill_type='solid')
        ws['J'+nextbegin].fill = fillinfo
        ws["J"+nextbegin] = "晚"
        ws["J"+nextbegin].font = font
        ws['J'+nextbegin].alignment = align

        # ws["K"+nextbegin] = "P-045"
        # ws["K"+nextbegin].font = font1
        # ws['K'+nextbegin].alignment = align


        ws["M"+nextbegin] = "脱模时间"
        ws["M"+nextbegin].font = font
        ws['M'+nextbegin].alignment = align

        nextbegin = str(onepagebegin+3)
        ws["A"+nextbegin] = "特殊要求"
        ws["A"+nextbegin].font = font
        ws['A'+nextbegin].alignment = align
        
        tmpcell = 'B'+nextbegin+":L"+nextbegin
        ws.merge_cells(tmpcell)

        return onepagebegin+3

    def DisMainPartInfo(self, ws, onepagebegin, arralldata, arrsamemainpart):
        self.DisMainPartHeader(ws, onepagebegin)
        # for key in arrwritedata:
        #     arrvalue = arrwritedata[key]
        #     # erplogger.info(f"key={key}")
        # mainnum = onerowdata['mainnum']
        # for data in dictsamemainpart[mainnum]:
        # for data in dictsamemainpart[mainnum]:
        for data in arrsamemainpart:
            onepagebegin=onepagebegin+1
            # valueinfo["mainnum"]
            # mainnum = data["mainnum"]
            # maihavesubpart = dictmaihavesubpart[mainnum]
            # onecnt = 0.0;
            # totalcnt = 0.0;
            # for mainpart in maihavesubpart:
            #     if mainpart["subpartnum"] in arrsamemainpartkey:
            #         print("mainpart=", mainpart)
            #         onecnt = onecnt + float(mainpart["onecnt"]);
            #         totalcnt = onecnt + float(mainpart["totalcnt"]);

            # data["onecnt"] = onecnt
            # data["totalcnt"] = totalcnt
            dateinfo = ""
            for tmpdata in arralldata:
                if tmpdata['mainnum'] == data:
                    dateinfo = tmpdata['date'];
                    self.DisMainPartBody(ws, onepagebegin, tmpdata)
                    break
        # for i in range(maincnt):
        #     onepagebegin=onepagebegin+1
        #     self.DisMainPartBody(ws, onepagebegin)
        # onepagebegin=onepagebegin+1
        # self.DisMainPartBody(ws, onepagebegin)
        # onepagebegin=onepagebegin+1
        # self.DisMainPartBody(ws, onepagebegin)
        # onepagebegin=onepagebegin+1
        # self.DisMainPartBody(ws, onepagebegin)
        # onepagebegin=onepagebegin+1
        # self.DisMainPartBody(ws, onepagebegin)
        # onepagebegin=onepagebegin+1
        # self.DisMainPartBody(ws, onepagebegin)
        return onepagebegin, dateinfo

    def DisBolder(self, ws, begininfo, endinfo):
        border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
        
        # if isfirst == 1:
        for i in range(begininfo, endinfo+1):
                for row in ws.iter_rows(min_row=i, max_row=i, min_col=1, max_col=13):
                    for cell in row:
                        cell.border = border

    def disend(self, ws, onepagebegin, showtimeinfo):
        align = Alignment(horizontal='center', vertical='center')
        font = Font(color=colors.BLACK, bold=True, size=10)
        font1 = Font(size = 9)
        nextbegin = str(onepagebegin)
        ws["A"+nextbegin] = "日期:"
        ws["A"+nextbegin].font = font1
        ws['A'+nextbegin].alignment = align

        ws["B"+nextbegin] = showtimeinfo
        ws["B"+nextbegin].font = font1
        ws['B'+nextbegin].alignment = align

        ws["F"+nextbegin] = "计算:"
        ws["F"+nextbegin].font = font1
        ws['F'+nextbegin].alignment = align

        # ws["G"+nextbegin] = "噜噜噜"
        ws["G"+nextbegin] = ""
        ws["G"+nextbegin].font = font1
        ws['G'+nextbegin].alignment = align

        ws["K"+nextbegin] = "排产:"
        ws["K"+nextbegin].font = font1
        ws['K'+nextbegin].alignment = align

        # ws["L"+nextbegin] = "噜噜噜收到"
        ws["L"+nextbegin] = ""
        ws["L"+nextbegin].font = font1
        ws['L'+nextbegin].alignment = align

    def onepage(self, ws, onepagebegin, pagenum, arralldata, dictmaihavesubpart, dictsamemainpart):
       
        cnt=0
        for key in dictsamemainpart:
            if  cnt % 3 == 0:

                # onepagebegin = self.DisTaskOrderInfo(ws, onepagebegin, pagenum)
                if cnt == 0:
                    onepagebegin = self.DisTaskOrderInfo(ws, onepagebegin, pagenum)
                else:
                    onepagebegin = self.DisTaskOrderInfo(ws, onepagebegin+3, pagenum)
                begininfo = onepagebegin
                pagenum = pagenum + 1
                # onepagebegin = onepagebegin+1
                # self.disend(ws, onepagebegin, "")
            onepagebegin, dateinfo = self.DisMainPartInfo(ws, onepagebegin, arralldata, dictsamemainpart[key])
            onepagebegin = onepagebegin + 1
            # mainnum = onerowdata['mainnum']
            onepagebegin = self.DisSubPar(ws, onepagebegin, dictmaihavesubpart[key])
            onepagebegin = onepagebegin+1
            cnt = cnt + 1
            if cnt > 0 and cnt % 3 == 0:
                onepagebegin = onepagebegin+1
                self.disend(ws, onepagebegin, dateinfo)
            endinfo = onepagebegin
            self.DisBolder(ws, begininfo, endinfo)

        if cnt > 0 and cnt % 3 != 0:
                onepagebegin = onepagebegin+1
                self.disend(ws, onepagebegin, dateinfo)

        endinfo = onepagebegin
        self.DisBolder(ws, begininfo, endinfo)
        return onepagebegin + 1
     

        # self.disend(ws, onepagebegin)

    def thread_function(self):
        # 这里是线程要执行的代码
        
        timeinfo = self.dateTimeEdit.dateTime().toString('yyyyMMdd')
        
        # print("timeinfo11111=", timeinfo, " exitteam=", self.editTeam.text())
        showtimeinfo = self.dateTimeEdit.dateTime().toString('yyyy.MM.dd')
        
        # print("timeinfo222=", timeinfo, " exitteam=", self.editTeam.text())
        
        # return
        # print("timeinfo333=", timeinfo, " exitteam=", self.editTeam.text())
        # timeinfo = "20240103"
        # self.editTeam.setText("浇注乙班")
        # print("timeinfo=", timeinfo, " exitteam=", self.editTeam.text())
        # # 连接到SQL Server　tds_version='7.0'
        try:
            conn = pymssql.connect(server=setting.sqlhost, user=setting.sqluser, password=setting.sqlpassword, database=setting.sqldatabase, tds_version=setting.tds_version)
            
            # 创建一个cursor对象
            cursor = conn.cursor()
        #     sql = """
        #     select sa.RAA015, ta.DEA002, ta.DEA057,sa.RAA961,sa.RAA021,sa.RAA980,sa.RAA981,sa.RAA018,sab.RAB003,sab.RAB004,sab.RAB981,sab.RAB007,sab.RAB981, sab.RAB007 from SGMRAA sa left join TPADAA taa on sa.RAA005=taa.DAA001 left join TPADEA ta on sa.RAA015=ta.DEA001 join SGMRAB sab on sa.RAA001 = sab.RAB001 where sa.RAA006='{0}'
        # """
            # WHERE SGMRAA.RAA024 = 'T'  and SGMRAA.RAA020='N'
            sql = """
            WITH MainComponentInfo AS (
        SELECT
            TPADAA.DAA002 AS 班组,
            SGMRAA.RAA015 AS 主件品号,
            TPADEA.DEA002 AS 主件品名,
            TPADEA.DEA057 AS 主件规格,
            SGMRAA.RAA961 AS 模具尺寸,
            SGMRAA.RAA021 AS 单头备注,
            SGMRAA.RAA980 AS 模具数量,
            SGMRAA.RAA981 AS 生产次数,
            SGMRAA.RAA018 AS 生产数量,
            SGMRAB.RAB981 AS "单次用量(kg)",
            SGMRAB.RAB007 AS "合计用量(kg)",
            SGMRAB.RAB003 AS 子件品号,
            SGMRAB.RAB004 AS 子件品名,
            SGMRAB.RAB981*1000 AS "配料区(kg)",
            SGMRAA.RAA006 AS 日期,
            SGMRAA.RAA001 AS 工单单号,
            ROW_NUMBER() OVER (PARTITION BY SGMRAA.RAA001 ORDER BY SGMRAB.RAB002 ASC) AS 子件序号, -- 按 RAB002 升序排序
            CASE WHEN SGMRAA.RAA981 = 0 THEN NULL ELSE (SGMRAB.RAB007 / SGMRAA.RAA981) END AS 单次用量
        FROM SGMRAA
        LEFT JOIN TPADEA ON TPADEA.DEA001 = SGMRAA.RAA015
        LEFT JOIN TPADAA ON TPADAA.DAA001 = SGMRAA.RAA005
        LEFT JOIN SGMRAB ON SGMRAA.RAA001 = SGMRAB.RAB001
        WHERE SGMRAA.RAA024 = 'T' and SGMRAA.RAA020='N'
    ),
    PivotedComponentInfo AS (
        SELECT
            MCI.班组,
            MCI.主件品号,
            MCI.主件品名,
            MCI.主件规格,
            MCI.模具尺寸,
            MCI.单头备注,
            MCI.模具数量,
            MCI.生产次数,
            MCI.生产数量,
            SUM(MCI.单次用量) AS "单次用量总和",
            SUM(MCI."合计用量(kg)") AS "合计用量总和",
            MCI.日期,
            MCI.工单单号,
            STUFF((SELECT ',' + 子件品号
                FROM MainComponentInfo MCI2
                WHERE MCI2.工单单号 = MCI.工单单号 AND MCI2.日期 = MCI.日期
                FOR XML PATH('')), 1, 1, '') AS 材料组合1,
            MAX(CASE WHEN MCI.子件序号 = 1 THEN MCI.子件品号 ELSE NULL END) AS 子件品号1,
            MAX(CASE WHEN MCI.子件序号 = 2 THEN MCI.子件品号 ELSE NULL END) AS 子件品号2,
            MAX(CASE WHEN MCI.子件序号 = 3 THEN MCI.子件品号 ELSE NULL END) AS 子件品号3,
            MAX(CASE WHEN MCI.子件序号 = 4 THEN MCI.子件品号 ELSE NULL END) AS 子件品号4,
            MAX(CASE WHEN MCI.子件序号 = 5 THEN MCI.子件品号 ELSE NULL END) AS 子件品号5,
            MAX(CASE WHEN MCI.子件序号 = 6 THEN MCI.子件品号 ELSE NULL END) AS 子件品号6,
            MAX(CASE WHEN MCI.子件序号 = 7 THEN MCI.子件品号 ELSE NULL END) AS 子件品号7,
            MAX(CASE WHEN MCI.子件序号 = 1 THEN MCI.子件品名 ELSE NULL END) AS 子件品名1,
            MAX(CASE WHEN MCI.子件序号 = 2 THEN MCI.子件品名 ELSE NULL END) AS 子件品名2,
            MAX(CASE WHEN MCI.子件序号 = 3 THEN MCI.子件品名 ELSE NULL END) AS 子件品名3,
            MAX(CASE WHEN MCI.子件序号 = 4 THEN MCI.子件品名 ELSE NULL END) AS 子件品名4,
            MAX(CASE WHEN MCI.子件序号 = 5 THEN MCI.子件品名 ELSE NULL END) AS 子件品名5,
            MAX(CASE WHEN MCI.子件序号 = 6 THEN MCI.子件品名 ELSE NULL END) AS 子件品名6,
            MAX(CASE WHEN MCI.子件序号 = 7 THEN MCI.子件品名 ELSE NULL END) AS 子件品名7,
            MAX(CASE WHEN MCI.子件序号 = 1 THEN MCI.单次用量 ELSE NULL END) AS 单次用量1,
            MAX(CASE WHEN MCI.子件序号 = 2 THEN MCI.单次用量 ELSE NULL END) AS 单次用量2,
            MAX(CASE WHEN MCI.子件序号 = 3 THEN MCI.单次用量 ELSE NULL END) AS 单次用量3,
            MAX(CASE WHEN MCI.子件序号 = 4 THEN MCI.单次用量 ELSE NULL END) AS 单次用量4,
            MAX(CASE WHEN MCI.子件序号 = 5 THEN MCI.单次用量 ELSE NULL END) AS 单次用量5,
            MAX(CASE WHEN MCI.子件序号 = 6 THEN MCI.单次用量 ELSE NULL END) AS 单次用量6,
            MAX(CASE WHEN MCI.子件序号 = 7 THEN MCI.单次用量 ELSE NULL END) AS 单次用量7
            
        FROM MainComponentInfo MCI
        WHERE MCI.班组 LIKE '%{0}%'  AND MCI.日期='{1}'
        GROUP BY
            MCI.班组,
            MCI.主件品号,
            MCI.主件品名,
            MCI.主件规格,
            MCI.模具尺寸,
            MCI.单头备注,
            MCI.模具数量,
            MCI.生产次数,
            MCI.生产数量,
            MCI.日期,
            MCI.工单单号
    )

    SELECT
        主件品号,
        主件品名,
        主件规格,
        模具尺寸,
        单头备注,
        模具数量,
        生产数量,
        生产次数,
        "单次用量总和",
        "合计用量总和",
        子件品号1,
        子件品号2,
        子件品号3,
        子件品号4,
        子件品号5,
        子件品号6,
        子件品号7,
        子件品名1,
        子件品名2,
        子件品名3,
        子件品名4,
        子件品名5,
        子件品名6,
        子件品名7,
        单次用量1,
        单次用量2,
        单次用量3,
        单次用量4,
        单次用量5,
        单次用量6,
        单次用量7,
        材料组合1,
        日期
    FROM PivotedComponentInfo;
        """
            # sql = sql.format(timeinfo)
            # if self.editTeam.text() != "":
            #     sql = sql + """  and taa.DAA002 LIKE '%{0}%'"""
            sql = sql.format(self.editTeam.text(), timeinfo)
            # print("sql=", sql)
            erplogger.info(f"sql={sql}")
            cursor.execute(sql)
            arralldata = []
            dictalldata = {}
            
            rows = cursor.fetchall()
            # print("len=", len(rows))
            erplogger.info(f"len={len(rows)}")
            
            dictmaihavesubpart = {}
            for row in rows:
                # erplogger.info(f"rows={row}")
                maininfo = {
                    "mainnum":row[0],
                    "mainname":row[1].encode('latin1').decode('gbk'),
                    "mainspec":row[2].encode('latin1').decode('gbk'),
                    "mainsize":row[3].encode('latin1').decode('gbk'),
                    "mainnotes":row[4].encode('latin1').decode('gbk'),
                    "dienum":row[5],
                    "prodnum":row[6],
                    "prodcnt":row[7],
                    "onecnt":row[8],
                    "totalcnt":row[9],
                    "singledose1":row[24],
                    "singledose2":row[25],
                    "singledose3":row[26],
                    "singledose4":row[27],
                    "singledose5":row[28],
                    "singledose6":row[29],
                    "singledose7":row[30],
                    "date":row[32]
                }

                erplogger.info(f"mainname={maininfo['mainname']} mainspec={maininfo['mainspec']} mainsize={maininfo['mainsize']} mainnotes={maininfo['mainnotes']}  date={maininfo['date']}")

                arrsubpart = []
                for i in range(10, 17):
                    if row[i] is not None:
                        subpart = {
                            "subpartnum":row[i],
                            "subpartname":row[i+7].encode('latin1').decode('gbk'),
                            "IngredientsArea":0.0
                        }

                        # erplogger.info(f"subpartnum={subpart['subpartnum']} subpartname={subpart['subpartname']}")
                        arrsubpart.append(subpart)

                dictmaihavesubpart[row[0]] = arrsubpart
                arralldata.append(maininfo)
                dictalldata[row[0]] = maininfo

        except pymssql.Error as e:
            # 捕获PyMSSQL异常
            erplogger.info(f"PyMSSQL Error: {e}")
            return
       
        copydictmaihavesubpart = dictmaihavesubpart.copy()
        dictsamemainpart = {}
        while True:
            if len(copydictmaihavesubpart) <= 0:
                break
            # print("prc")
            first_element = next(iter(copydictmaihavesubpart))
            arrvalue = copydictmaihavesubpart[first_element]
            arrallsame = []
            for keysel in copydictmaihavesubpart:
                arrselvalue = copydictmaihavesubpart[keysel]
                if len(arrvalue) != len(arrselvalue):
                    continue

                currlen = len(arrvalue) 
                isallsame = True
                for i in range(currlen):
                    if arrvalue[i]["subpartnum"] != arrselvalue[i]["subpartnum"]:
                        isallsame = False
                        break;
            
                if isallsame == True:
                    arrallsame.append(keysel)

            dictsamemainpart[first_element] = arrallsame  
            for key in arrallsame:
                del copydictmaihavesubpart[key]

        for key in dictmaihavesubpart:
            arrsubpart = dictmaihavesubpart[key]
            leninfo = len(arrsubpart)
            if key in dictsamemainpart:
                dicttmp = {}
                for i in range(leninfo):
                    dicttmp["val"+str(i)] = 0.0
                arrmainpart = dictsamemainpart[key]
                
                for mainpart in arrmainpart:
                    # print("mainpart=", mainpart)
                    for i in range(0, leninfo):
                        # print("singledose=", dictalldata[mainpart]['singledose'+str(i+1)], " subpart=", arrsubpart[i])
                        dicttmp["val"+str(i)] = dicttmp["val"+str(i)] + float(dictalldata[mainpart]['singledose'+str(i+1)])

                for i in range(leninfo):
                    arrsubpart[i]['IngredientsArea'] = dicttmp["val"+str(i)]

        # dictsamemainpart = {}
        # for key in dictmaihavesubpart:
        #     arrvalue = dictmaihavesubpart[key]
        #     # erplogger.info(f"key={key}")
        #     # for data in arrvalue:
        #         # erplogger.info(f"mainnum={data['mainnum']}, mainname={data['mainname']}, mainspec={data['mainspec']}, mainsize={data['mainsize']}, mainnotes={data['mainnotes']}, dienum={data['dienum']}, prodnum={data['prodnum']}, prodcnt={data['prodcnt']}, subpartnum={data['subpartnum']}, subpartname={data['subpartname']}")
            
        #     arrallsame = []
        #     for keysel in dictmaihavesubpart:
        #         arrselvalue = dictmaihavesubpart[keysel]
        #         if len(arrvalue) != len(arrselvalue):
        #             continue
                
        #         currlen = len(arrvalue) 
        #         isallsame = True
        #         for i in range(currlen):
        #             if arrvalue[i]["subpartnum"] != arrselvalue[i]["subpartnum"]:
        #                 isallsame = False
        #                 break;
        #         if isallsame == True:
        #             # arrallsame.append(keysel)
        #             # tmpdata = {
        #             #     "mainpartnum":keysel
        #             # }
        #             arrallsame.append(keysel)

        #     dictsamemainpart[key] = arrallsame  

        # return
        # for data in arralldata:
            # erplogger.info(f"mainnum={data['mainnum']}, mainname={data['mainname']}, mainspec={data['mainspec']}, mainsize={data['mainsize']}, mainnotes={data['mainnotes']}, dienum={data['dienum']}, prodnum={data['prodnum']}, prodcnt={data['prodcnt']}, subpartnum={data['subpartnum']}, subpartname={data['subpartname']}")
        cursor.close()
        conn.close()
        filepath = self.editFilePath.text()
        print("exportfile: func begin text=", filepath)
        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active # 获取当前活动的工作表，默认情况下，活动工作表是Excel文件中的第一个工作表，也可以通 
                    # 过wb.active = index来设置活动工作表的索引。
        
        ws.title = "sheet";
        self.setsheetwidth(ws)
        onepagebegin = 1
        pagenum = 1
        onepagebegin = self.onepage(ws, onepagebegin, pagenum, arralldata, dictmaihavesubpart, dictsamemainpart)
        # pagenum = pagenum + 1
        # onepagebegin = self.onepage(ws, onepagebegin, pagenum, arralldata)
        wb.save(filepath)
        self.exporthint.setText("导出数据到文件成功")
        # QMessageBox.information(self, '提示信息', '导出数据到文件成功')  # 可以不带选择按键，也可以带ok按键
    def exportfile(self):
        self.exporthint.setText("正在导出文件  请稍等")
       
        # 关闭连接
        
        # return
        # arrsubpart = []
        # cursor.execute("select saa.RAA015,sab.RAB003,sab.RAB004 from SGMRAA saa join SGMRAB sab on saa.RAA001 = sab.RAB001")
        # rows = cursor.fetchall()
        # for row in rows:
        #     subpart = {
        #         "mainpartnum":row[0],
        #         "subpartnum":row[1],
        #         "subpartname":row[2]
        #     }

        #     arrsubpart.append(subpart)


        # for subdata in arrsubpart:
        #     for maindata in mainpart.arrmainpart: 
        #         if subdata['mainpartnum'] == maindata['mainnum']:
        #             print("mainpart=", maindata['mainnum'], " subpartnum=", subdata['subpartnum'], " subpartname=", subdata['subpartname'])

        #     # print("mainnum=", data['mainnum'])
        # # 关闭cursor
        # cursor.close()
        
        # # 关闭连接
        # conn.close()
        # return
        # print("111222")
        # self.editTeam.setText("浇注乙班")
        # 创建线程
        my_thread = threading.Thread(target=self.thread_function)
        # 设置为守护线程
        # my_thread.daemon = True
        my_thread.start()
     
        # ws.merge_cells(start_row=1, end_row=3, start_column=1, end_column=1)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=2, end_column=2)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=3, end_column=3)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=4, end_column=4)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=5, end_column=5)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=6, end_column=6)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=7, end_column=7)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=8, end_column=8)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=9, end_column=9)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=10, end_column=10)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=11, end_column=11)#标底合并
        # ws.merge_cells(start_row=1, end_row=3, start_column=12, end_column=12)#标底合并
        # 保存工作簿到文件
       
        

if __name__ == '__main__':
    app = QApplication(sys.argv)
    vieo_gui = sportMainWindow()
    vieo_gui.show()
    sys.exit(app.exec_())


